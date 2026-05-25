# src/data_engine/extractors/bancolombia.py
import polars as pl
import csv
import logging
import re
import openpyxl
from .base import BaseExtractor

logger = logging.getLogger("bancolombia_extractor")

class BancolombiaExtractor(BaseExtractor):
    def process(self) -> pl.DataFrame:
        try:
            data = []
            
            # 1. Función unificada para esterilizar el valor
            def limpiar_valor(val_str):
                if val_str is None or val_str == "":
                    return None
                v = str(val_str).strip()
                v = re.sub(r'[\$\s\xa0]', '', v)
                try:
                    return float(v)
                except ValueError:
                    return None

            # 2. LECTURA DEL ARCHIVO: Cascada de Intentos
            try:
                # Intento Principal: Leer como Excel (.xlsx)
                wb = openpyxl.load_workbook(self.filepath, data_only=True)
                
                # --- FORMATO A: Lectura de la primera hoja (sin importar su nombre) ---
                ws_mov = wb[wb.sheetnames[0]]
                headers = [str(c.value).strip().upper() if c.value else "" for c in ws_mov[1]]
                
                if "FECHA" in headers and "CONCEPTO" in headers and "VALOR" in headers:
                    idx_f = headers.index("FECHA")
                    idx_c = headers.index("CONCEPTO")
                    idx_v = headers.index("VALOR")
                    
                    for row in ws_mov.iter_rows(min_row=2):
                        fecha_val = row[idx_f].value
                        concepto_val = row[idx_c].value
                        celda_valor = row[idx_v]
                        
                        if not fecha_val or not concepto_val or celda_valor.value is None:
                            continue
                            
                        valor = limpiar_valor(celda_valor.value)
                        if valor is None:
                            continue

                        # Detección del color de fuente (La lógica clásica)
                        es_rojo = False
                        if celda_valor.font and celda_valor.font.color:
                            color_val = getattr(celda_valor.font.color, "rgb", None)
                            if color_val and "FF0000" in str(color_val).upper():
                                es_rojo = True
                        
                        if valor < 0 or es_rojo:
                            ingreso = 0.0
                            egreso = abs(valor)
                        else:
                            ingreso = abs(valor)
                            egreso = 0.0
                            
                        data.append({
                            "Fecha": str(fecha_val),
                            "Concepto": str(concepto_val),
                            "Documento_Referencia": "N/A",
                            "Ingreso": ingreso,
                            "Egreso": egreso,
                            "Origen": "BANCOLOMBIA"
                        })

                # --- FORMATO B y C: Los formatos nuevos (Si el antiguo no extrajo nada) ---
                if not data:
                    ws = wb.active
                    primera_celda = ws.cell(1, 1).value
                    
                    if primera_celda and ',' in str(primera_celda):
                        # FORMATO B: Formato nuevo (Todo en columna A separado por comas)
                        for row in ws.iter_rows(values_only=True):
                            if row[0] is None:
                                continue
                            campos = str(row[0]).split(',')
                            if len(campos) < 8:
                                continue
                                
                            raw_fecha = campos[3].strip() if len(campos) > 3 else ""
                            raw_valor = campos[5].strip() if len(campos) > 5 else ""
                            concepto = campos[7].strip() if len(campos) > 7 else ""
                            
                            if not raw_fecha or not raw_valor or not concepto:
                                continue
                                
                            valor = limpiar_valor(raw_valor)
                            if valor is None:
                                continue
                            
                            if valor > 0:
                                ingreso = valor
                                egreso = 0.0
                            else:
                                ingreso = 0.0
                                egreso = abs(valor)
                                
                            data.append({
                                "Fecha": raw_fecha,
                                "Concepto": concepto,
                                "Documento_Referencia": "N/A",
                                "Ingreso": ingreso,
                                "Egreso": egreso,
                                "Origen": "BANCOLOMBIA"
                            })
                    else:
                        # FORMATO C: Formato original nuevo (Múltiples columnas fijas)
                        for row in ws.iter_rows(values_only=True):
                            if len(row) < 8:
                                continue
                            raw_fecha = str(row[3]).strip() if row[3] else ""
                            raw_valor = str(row[5]).strip() if row[5] else ""
                            concepto = str(row[7]).strip() if row[7] else ""
                            
                            if not raw_fecha or not raw_valor or not concepto:
                                continue
                                
                            valor = limpiar_valor(raw_valor)
                            if valor is None:
                                continue
                            
                            if valor > 0:
                                ingreso = valor
                                egreso = 0.0
                            else:
                                ingreso = 0.0
                                egreso = abs(valor)
                                
                            data.append({
                                "Fecha": raw_fecha,
                                "Concepto": concepto,
                                "Documento_Referencia": "N/A",
                                "Ingreso": ingreso,
                                "Egreso": egreso,
                                "Origen": "BANCOLOMBIA"
                            })
                wb.close()
                
            except Exception as e:
                # --- FORMATO D: Fallback de emergencia a texto/CSV puro ---
                logger.warning(f"No se pudo leer como Excel, intentando como CSV puro: {e}")
                with open(self.filepath, mode='r', encoding='utf-8', errors='replace') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) < 8:
                            continue
                        raw_fecha = row[3].strip()
                        raw_valor = row[5].strip()
                        concepto = row[7].strip()
                        
                        if not raw_fecha or not raw_valor or not concepto:
                            continue
                        
                        valor = limpiar_valor(raw_valor)
                        if valor is None:
                            continue
                        
                        if valor > 0:
                            ingreso = valor
                            egreso = 0.0
                        else:
                            ingreso = 0.0
                            egreso = abs(valor)
                            
                        data.append({
                            "Fecha": raw_fecha,
                            "Concepto": concepto,
                            "Documento_Referencia": "N/A",
                            "Ingreso": ingreso,
                            "Egreso": egreso,
                            "Origen": "BANCOLOMBIA"
                        })

            if not data:
                return pl.DataFrame()

            # 3. TRANSFORMACIÓN UNIFICADA A POLARS
            # Esta limpieza de fecha funciona tanto para el YYYY-MM-DD HH:MM:SS del Excel clásico 
            # como para las fechas raras del CSV
            df_clean = pl.DataFrame(data).with_columns([
                pl.col("Fecha")
                  .str.replace_all(r"[-/:\s]", "")
                  .str.slice(0, 8)
                  .str.to_date("%Y%m%d", strict=False)
            ]).filter(pl.col("Fecha").is_not_null())

            # 4. REGLAS DE NEGOCIO GLOBALES
            df_final = df_clean.with_columns(
                pl.when(pl.col("Concepto").str.contains(r"(?i)TRASL ENTRE FONDOS DE VALORES"))
                .then(pl.lit("Traslado_Salida"))
                
                .when(pl.col("Concepto").str.contains(r"(?i)PAGO DE PROV CCA ALIANZA FID"))
                .then(pl.lit("Traslado_Entrada"))
                
                .otherwise(pl.lit("Operacion_Normal"))
                .alias("Categoria_Flujo")
            )

            print(f"[OK] Bancolombia: {len(df_final)} movimientos extraidos")
            return df_final
            
        except Exception as e:
            logger.error(f"Error procesando Bancolombia ({self.filepath}): {e}")
            # Retornamos la estructura base para no romper el consolidado
            return pl.DataFrame({"Fecha": [], "Concepto": [], "Documento_Referencia": [], "Ingreso": [], "Egreso": [], "Origen": [], "Categoria_Flujo": []})